# -*- coding: utf-8 -*-
# 第一行必须有，否则报中文字符非ascii码错误

import urllib.request
from urllib.parse import quote
import string
import json
import time
# import xlwt
from openpyxl import load_workbook
import openpyxl


#初始化xlsx表头
def initExcel(path):
    book = openpyxl.Workbook()  # 新建一个excel
    sheet = book.active  # 添加一个sheet页
    sheet['A1'] = 'name'
    sheet['B1'] ='lat'
    sheet['C1'] = 'log'
    sheet['D1'] ='address'
    sheet['E1'] ='province'
    sheet['F1'] ='city'
    sheet['G1'] ='area'
    sheet['H1'] ='detail'
    sheet['I1'] = 'uid'
    book.save(path)

#数据导入xlsx
def importData(path,i):
    book = openpyxl.load_workbook(path)
    worksheet=book.active
    row=worksheet.max_row+1
    worksheet.cell(row,1,i['name'])
    worksheet.cell(row,2,i['location']['lat'])
    worksheet.cell(row, 3, i['location']['lng'])
    worksheet.cell(row, 4, i['address'])
    worksheet.cell(row, 5, i['province'])
    worksheet.cell(row, 6, i['city'])
    worksheet.cell(row, 7, i['area'])
    worksheet.cell(row, 8, i['detail'])
    worksheet.cell(row, 9, i['uid'])
    book.save(path)



# ak需要在百度地图开放平台申请
ak="7XQ17G7GYUIrz70leZXK5nhSzM1xbtx4"

# 关键词
query = ["停车场"]
page_size = 20
page_num = 0
scope = 1

# 范围：
# 左下坐标 30.379,114.118
# 右上坐标 30.703,114.665
# 中间坐标 30.541,114.3915

bounds = [
    [30.379, 114.118, 30.541, 114.3915],
    [30.379, 114.3915, 30.541, 114.665],
    [30.541, 114.118, 30.703, 114.3915],
    [30.541, 114.3915, 30.703, 114.665]
]

new_bounds = []
# col_row 将bounds的每一小块继续细分为3行3列，可以防止区域内的搜索数量上限400
col_row = 3
for lst in bounds:
    distance_lat = (lst[2] - lst[0]) / col_row
    distance_lon = (lst[3] - lst[1]) / col_row
    for i in range(col_row):
        for j in range(col_row):
            lst_temp = []
            lst_temp.append(lst[0] + distance_lat * i)
            lst_temp.append(lst[1] + distance_lon * j)
            lst_temp.append(lst[0] + distance_lat * (i + 1))
            lst_temp.append(lst[1] + distance_lon * (j + 1))
            new_bounds.append(lst_temp)

initExcel('D:/test.xlsx')

for bound in new_bounds:
    np = True
    a = []
    while np == True:
        # 使用百度提供的url拼接条件
        url = "http://api.map.baidu.com/place/v2/search?ak=" + str(ak) + "&output=json&query=" + str(
            query[0]) + "&page_size=" + str(page_size) + "&page_num=" + str(page_num) + "&bounds=" + str(
            bound[0]) + "," + str(bound[1]) + "," + str(bound[2]) + "," + str(bound[3])
        url = quote(url, safe=string.printable)

        # 请求url读取，创建网页对象
        jsonf = urllib.request.urlopen(url)
        page_num = page_num + 1
        # 判断查询翻页进程
        jsonfile = jsonf.read()
        s = json.loads(jsonfile)
        total = int(s["total"])
        a.append(total)
        #导入Excel
        if len(s['results']) != 0:
            for i in s['results']:
                if i!=None:
                    importData('D:/test.xlsx',i)


        max_page = int(a[0] / page_size) + 1
        # 防止并发过高，百度地图要求并发小于120
        time.sleep(1)

        if page_num > max_page:
            np = False
            page_num = 0
            # print("search complete")
            # print("output: " + str(bound))
            # print("total: " + str(a[0]))
            # print("")

# results = open("D:/results.json", 'a')
# results.write(str(queryResults))
# results.close()
print("ALL DONE!")
