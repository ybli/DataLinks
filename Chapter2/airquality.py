# -*- coding: utf-8 -*-

import urllib.request
from urllib.parse import quote
import string
import json
import time
from openpyxl import load_workbook
import openpyxl
import threading
from queue import Queue

#初始化xlsx表头
def initExcel(path):
    book = openpyxl.Workbook()  # 新建一个excel
    sheet = book.active  # 添加一个sheet页
    sheet['A1'] = 'xh'
    sheet['B1'] ='jcsj'
    sheet['C1'] = 'jcdwmc'
    sheet['D1'] ='zslb'
    sheet['E1'] ='aqi'
    sheet['F1'] ='zsjb'
    sheet['G1'] ='ysjb'
    sheet['H1'] ='sywrw'
    sheet['I1'] = 'jkyxzk'
    book.save(path)

#数据导入xlsx
def importData(path,i):
    book = openpyxl.load_workbook(path)
    worksheet=book.active
    row=worksheet.max_row+1

    # 1
    if "xh" in i:
        worksheet.cell(row, 1, i['xh'])
    else:
        worksheet.cell(row, 1, "")
    # 2
    if "jcsj" in i:
        worksheet.cell(row, 2, i['jcsj'])
    else:
        worksheet.cell(row, 2, "")
    # 3
    if "jcdwmc" in i:
        worksheet.cell(row, 3, i['jcdwmc'])
    else:
        worksheet.cell(row, 3, "")
    # 4
    if "zslb" in i:
        worksheet.cell(row, 4, i['zslb'])
    else:
        worksheet.cell(row, 4, "")
    # 5
    if "aqi" in i:
        worksheet.cell(row, 5, i['aqi'])
    else:
        worksheet.cell(row, 5, "")
    # 6
    if "zsjb" in i:
        worksheet.cell(row, 6, i['zsjb'])
    else:
        worksheet.cell(row, 6, "")
    # 7
    if "ysjb" in i:
        worksheet.cell(row, 7, i['ysjb'])
    else:
        worksheet.cell(row, 7, "")
    # 8
    if "sywrw" in i:
        worksheet.cell(row, 8, i['sywrw'])
    else:
        worksheet.cell(row, 8, "")
    # 9
    if "jkyxzk" in i:
        worksheet.cell(row, 9, i['jkyxzk'])
    else:
        worksheet.cell(row, 9, "")

    book.save(path)

initExcel('D:/test.xlsx')

a = []
page_size = 20

url = "https://opendata.sz.gov.cn/api/1920606096/1/service.xhtml?page=1&rows=100&appKey=8a0d412d579f412c84153182707d2c74"
url = quote(url, safe=string.printable)
jsonf = urllib.request.urlopen(url)
jsonfile = jsonf.read()
s = json.loads(jsonfile)
total = int(s["total"])
a.append(total)
#导入Excel
if len(s['data']) != 0:
    for i in s['data']:
        if i!=None:
            importData('D:/test.xlsx',i)

print("Done")
